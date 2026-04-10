from flask import Blueprint, render_template, request, session, jsonify, send_file
from app.database import get_db, log_action
from app.auth import role_required
import io
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

manager = Blueprint('manager', __name__, url_prefix='/manager')


@manager.route('/')
@role_required('admin', 'manager')
def index():
    return render_template('manager.html',
                           full_name=session['user_full_name'],
                           role=session['user_role'])


@manager.route('/api/employees')
@role_required('admin', 'manager')
def get_employees():
    db   = get_db()
    rows = db.execute(
        "SELECT login, first_name, last_name, patronymic, role FROM Employees WHERE role = 'employee' ORDER BY last_name"
    ).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


@manager.route('/api/overview')
@role_required('admin', 'manager')
def overview():
    db   = get_db()
    rows = db.execute("""
        SELECT wt.*, e.first_name, e.last_name
        FROM WeeklyTemplate wt
        JOIN Employees e ON wt.login = e.login
        WHERE wt.status = 'approved' AND e.role = 'employee'
        ORDER BY e.last_name, wt.weekday
    """).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


# ── Pending template approvals ────────────────────────────────────────────────

@manager.route('/api/pending-templates')
@role_required('admin', 'manager')
def pending_templates():
    db   = get_db()
    rows = db.execute("""
        SELECT wt.*, e.first_name, e.last_name
        FROM WeeklyTemplate wt
        JOIN Employees e ON wt.login = e.login
        WHERE wt.status != 'approved'
        ORDER BY e.last_name, wt.weekday
    """).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


@manager.route('/api/templates/<int:tid>/approve', methods=['POST'])
@role_required('admin', 'manager')
def approve_template(tid):
    db  = get_db()
    row = db.execute("SELECT * FROM WeeklyTemplate WHERE id=?", (tid,)).fetchone()
    if not row:
        db.close()
        return jsonify({'error': 'Не найдено'}), 404
    db.execute("UPDATE WeeklyTemplate SET status='approved', manager_note='' WHERE id=?", (tid,))
    log_action(db, session['user_login'], 'TEMPLATE_APPROVE',
               f"id={tid} login={row['login']} weekday={row['weekday']}")
    db.commit()
    db.close()
    return jsonify({'ok': True})


@manager.route('/api/templates/<int:tid>/reject', methods=['POST'])
@role_required('admin', 'manager')
def reject_template(tid):
    note = (request.get_json() or {}).get('note', '').strip()
    if not note:
        return jsonify({'error': 'Укажите причину отклонения'}), 400
    db  = get_db()
    row = db.execute("SELECT * FROM WeeklyTemplate WHERE id=?", (tid,)).fetchone()
    if not row:
        db.close()
        return jsonify({'error': 'Не найдено'}), 404
    db.execute("UPDATE WeeklyTemplate SET status='rejected', manager_note=? WHERE id=?",
               (note, tid))
    log_action(db, session['user_login'], 'TEMPLATE_REJECT',
               f"id={tid} login={row['login']} reason={note}")
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ── Change requests ───────────────────────────────────────────────────────────

@manager.route('/api/change-requests')
@role_required('admin', 'manager')
def list_change_requests():
    db   = get_db()
    rows = db.execute("""
        SELECT cr.*, e.first_name, e.last_name
        FROM ChangeRequests cr
        JOIN Employees e ON cr.login = e.login
        ORDER BY cr.created_at DESC
    """).fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])


@manager.route('/api/change-requests/<int:crid>/approve', methods=['POST'])
@role_required('admin', 'manager')
def approve_change_request(crid):
    db  = get_db()
    row = db.execute("SELECT * FROM ChangeRequests WHERE id=?", (crid,)).fetchone()
    if not row:
        db.close()
        return jsonify({'error': 'Не найдено'}), 404
    db.execute("""
        INSERT INTO ScheduleOverrides (login, date, time_start, time_end, note)
        VALUES (?,?,?,?,?)
        ON CONFLICT(login, date) DO UPDATE SET
          time_start=excluded.time_start, time_end=excluded.time_end, note=excluded.note
    """, (row['login'], row['target_date'], row['new_start'], row['new_end'],
          'Из запроса на изменение'))
    db.execute("UPDATE ChangeRequests SET status='approved', manager_note='' WHERE id=?", (crid,))
    log_action(db, session['user_login'], 'CHANGE_REQ_APPROVE',
               f"id={crid} login={row['login']} date={row['target_date']}")
    db.commit()
    db.close()
    return jsonify({'ok': True})


@manager.route('/api/change-requests/<int:crid>/reject', methods=['POST'])
@role_required('admin', 'manager')
def reject_change_request(crid):
    note = (request.get_json() or {}).get('note', '').strip()
    if not note:
        return jsonify({'error': 'Укажите причину отклонения'}), 400
    db  = get_db()
    row = db.execute("SELECT * FROM ChangeRequests WHERE id=?", (crid,)).fetchone()
    if not row:
        db.close()
        return jsonify({'error': 'Не найдено'}), 404
    db.execute("UPDATE ChangeRequests SET status='rejected', manager_note=? WHERE id=?",
               (note, crid))
    log_action(db, session['user_login'], 'CHANGE_REQ_REJECT',
               f"id={crid} login={row['login']} reason={note}")
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ── Meetings ──────────────────────────────────────────────────────────────────

@manager.route('/api/meetings', methods=['GET'])
@role_required('admin', 'manager')
def list_meetings():
    db   = get_db()
    rows = db.execute("SELECT * FROM Meetings ORDER BY date DESC, time_start").fetchall()
    result = []
    for r in rows:
        m     = dict(r)
        parts = db.execute(
            "SELECT login FROM MeetingParticipants WHERE meeting_id=?", (m['id'],)
        ).fetchall()
        m['participants'] = [p['login'] for p in parts]
        result.append(m)
    db.close()
    return jsonify(result)


@manager.route('/api/meetings', methods=['POST'])
@role_required('admin', 'manager')
def create_meeting():
    data  = request.get_json() or {}
    title = (data.get('title') or '').strip()
    desc  = (data.get('description') or '').strip()
    d     = (data.get('date') or '').strip()
    ts    = (data.get('time_start') or '').strip()
    te    = (data.get('time_end') or '').strip()
    parts = data.get('participants', [])

    if not all([title, d, ts, te]):
        return jsonify({'error': 'Заполните обязательные поля'}), 400
    if ts >= te:
        return jsonify({'error': 'Время начала должно быть раньше конца'}), 400

    login = session['user_login']
    db    = get_db()
    cur   = db.execute(
        "INSERT INTO Meetings (title,description,date,time_start,time_end,created_by) "
        "VALUES (?,?,?,?,?,?)", (title, desc, d, ts, te, login)
    )
    mid = cur.lastrowid
    for p in parts:
        db.execute(
            "INSERT OR IGNORE INTO MeetingParticipants (meeting_id,login) VALUES (?,?)",
            (mid, p)
        )
    log_action(db, login, 'MEETING_CREATE',
               f"id={mid} '{title}' {d} {ts}-{te}")
    db.commit()
    row = db.execute("SELECT * FROM Meetings WHERE id=?", (mid,)).fetchone()
    db.close()
    return jsonify(dict(row)), 201


@manager.route('/api/meetings/<int:mid>', methods=['DELETE'])
@role_required('admin', 'manager')
def delete_meeting(mid):
    db  = get_db()
    row = db.execute("SELECT * FROM Meetings WHERE id=?", (mid,)).fetchone()
    if not row:
        db.close()
        return jsonify({'error': 'Не найдено'}), 404
    db.execute("DELETE FROM MeetingParticipants WHERE meeting_id=?", (mid,))
    db.execute("DELETE FROM Meetings WHERE id=?", (mid,))
    log_action(db, session['user_login'], 'MEETING_DELETE',
               f"id={mid} title={row['title']}")
    db.commit()
    db.close()
    return jsonify({'ok': True})


# ── Excel export ──────────────────────────────────────────────────────────────

@manager.route('/api/export-excel')
@role_required('admin', 'manager')
def export_excel():
    """
    Export employee schedules to Excel.
    Query params: date_from, date_to (YYYY-MM-DD).
    Defaults to current 2 weeks if not provided.
    """
    today   = date.today()
    monday  = today - timedelta(days=today.weekday())
    default_from = monday.isoformat()
    default_to   = (monday + timedelta(days=13)).isoformat()

    date_from = request.args.get('date_from', default_from)
    date_to   = request.args.get('date_to',   default_to)

    db = get_db()

    # Fetch all employees (role=employee only)
    employees = db.execute(
        "SELECT login, first_name, last_name, patronymic FROM Employees "
        "WHERE role='employee' ORDER BY last_name"
    ).fetchall()

    # Approved weekly templates
    templates = {}
    for row in db.execute(
        "SELECT login, weekday, time_start, time_end FROM WeeklyTemplate WHERE status='approved'"
    ).fetchall():
        templates.setdefault(row['login'], {})[row['weekday']] = row

    # Date-specific overrides
    overrides = {}
    for row in db.execute(
        "SELECT login, date, time_start, time_end FROM ScheduleOverrides "
        "WHERE date BETWEEN ? AND ?", (date_from, date_to)
    ).fetchall():
        overrides.setdefault(row['login'], {})[row['date']] = row

    # Meetings for the period
    meetings_by_date = {}
    for row in db.execute(
        "SELECT m.*, GROUP_CONCAT(mp.login) as part_logins "
        "FROM Meetings m LEFT JOIN MeetingParticipants mp ON mp.meeting_id = m.id "
        "WHERE m.date BETWEEN ? AND ? GROUP BY m.id", (date_from, date_to)
    ).fetchall():
        meetings_by_date.setdefault(row['date'], []).append(dict(row))

    db.close()

    # Build list of dates in range
    start = date.fromisoformat(date_from)
    end   = date.fromisoformat(date_to)
    dates = []
    cur   = start
    while cur <= end:
        dates.append(cur)
        cur += timedelta(days=1)

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = 'Расписание'

    # Styles
    MAGENTA  = 'FFE5006E'
    DARK     = 'FF1E1E1E'
    HEADER_BG= 'FF2A2A2A'
    GREEN_BG = 'FF1A3A2A'
    GREEN_FG = 'FF4CDE8A'
    MEET_BG  = 'FF3A1A2A'
    MEET_FG  = 'FFFF66AA'
    GRAY_FG  = 'FF888888'
    WHITE    = 'FFF0F0F0'

    bold_white  = Font(bold=True,  color=WHITE,   size=11)
    bold_mag    = Font(bold=True,  color=MAGENTA, size=11)
    normal_white= Font(bold=False, color=WHITE,   size=10)
    small_gray  = Font(bold=False, color=GRAY_FG, size=9)
    green_font  = Font(bold=True,  color=GREEN_FG, size=10)
    meet_font   = Font(bold=False, color=MEET_FG,  size=9)

    thin = Side(style='thin', color='FF333333')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    # ── Row 1: title ──
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=1 + len(dates))
    title_cell = ws.cell(row=1, column=1,
                         value=f'Расписание сотрудников: {date_from} — {date_to}')
    title_cell.font      = Font(bold=True, color=MAGENTA, size=14)
    title_cell.fill      = PatternFill('solid', fgColor=DARK)
    title_cell.alignment = left
    ws.row_dimensions[1].height = 28

    # ── Row 2: column headers ──
    DAYS_RU = ['Пн','Вт','Ср','Чт','Пт','Сб','Вс']
    ws.cell(row=2, column=1, value='Сотрудник').font = bold_white
    ws.cell(row=2, column=1).fill      = PatternFill('solid', fgColor=HEADER_BG)
    ws.cell(row=2, column=1).alignment = center
    ws.cell(row=2, column=1).border    = border
    ws.column_dimensions['A'].width    = 24
    ws.row_dimensions[2].height        = 30

    for i, d in enumerate(dates, start=2):
        col  = i + 1
        day_ru = DAYS_RU[d.weekday()]
        label  = f'{day_ru}\n{d.strftime("%d.%m")}'
        cell   = ws.cell(row=2, column=col, value=label)
        cell.font      = bold_white
        cell.fill      = PatternFill('solid', fgColor=HEADER_BG)
        cell.alignment = center
        cell.border    = border
        ws.column_dimensions[get_column_letter(col)].width = 14

    # ── Data rows ──
    for r_idx, emp in enumerate(employees, start=3):
        login     = emp['login']
        full_name = f"{emp['last_name']} {emp['first_name']}"

        name_cell = ws.cell(row=r_idx, column=1, value=full_name)
        name_cell.font      = normal_white
        name_cell.fill      = PatternFill('solid', fgColor=DARK)
        name_cell.alignment = left
        name_cell.border    = border
        ws.row_dimensions[r_idx].height = 36

        emp_tmpl = templates.get(login, {})
        emp_ovrd = overrides.get(login, {})

        for i, d in enumerate(dates, start=2):
            col     = i + 1
            iso     = d.isoformat()
            cell    = ws.cell(row=r_idx, column=col)
            cell.alignment = center
            cell.border    = border
            cell.fill      = PatternFill('solid', fgColor=DARK)

            # Determine shift
            shift = None
            is_override = False
            if iso in emp_ovrd:
                shift = emp_ovrd[iso]
                is_override = True
            elif d.weekday() in emp_tmpl:
                shift = emp_tmpl[d.weekday()]

            # Check for meetings on this date
            day_meetings = meetings_by_date.get(iso, [])
            emp_meetings = [
                m for m in day_meetings
                if not m['part_logins'] or login in (m['part_logins'] or '').split(',')
            ]

            lines = []
            if shift:
                ts = shift['time_start'][:5]
                te = shift['time_end'][:5]
                lines.append(f'{ts}–{te}')
                if is_override:
                    lines.append('(изм.)')
                cell.font = green_font
                cell.fill = PatternFill('solid', fgColor=GREEN_BG)

            for m in emp_meetings:
                mt = f"📋 {m['time_start'][:5]}–{m['time_end'][:5]}"
                lines.append(mt)
                if not shift:
                    cell.fill = PatternFill('solid', fgColor=MEET_BG)
                    cell.font = meet_font

            if lines:
                cell.value = '\n'.join(lines)
            else:
                cell.value = '—'
                cell.font  = small_gray

    # ── Freeze header ──
    ws.freeze_panes = 'B3'

    log_action(db if False else get_db(),
               session['user_login'], 'EXPORT_EXCEL',
               f"{date_from} to {date_to}")
    # (separate connection for log)
    _db = get_db()
    log_action(_db, session['user_login'], 'EXPORT_EXCEL', f"{date_from}—{date_to}")
    _db.commit()
    _db.close()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"schedule_{date_from}_{date_to}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
